import pdb
from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment
from openpyxl.styles import NamedStyle, Font, Border, Side


def get_xlsx_data(xlsx_file, sheet_name):

    FIELDS_ROW = 2
    FIRST_DATA_ROW = 7

    # verify that the xlsx file exists and has a sheet named 'Gear' if
    # anything goes wrong, print a message that says:

    # "something has gone wrong.  Please verify that the submitted
    # file is a valid and current GLFC Lamprey Template."

    # Load the workbook and select the active sheet
    wb = load_workbook(xlsx_file)

    sheet = wb[sheet_name]
    fields = [x.value.strip() for x in sheet[FIELDS_ROW] if x.value]
    data = []

    # Iterate through all rows
    for row in sheet.iter_rows(values_only=True, min_row=FIRST_DATA_ROW):
        if row.count(None) != len(row):
            record = {k: v for k, v in zip(fields[1:], row[1:])}
            data.append(record)

    return data


def write_error_report(errors, report_name, sheet_name):

    if report_name.exists():
        wb = load_workbook(report_name)
    else:
        wb = Workbook()

    ws = wb.create_sheet(sheet_name)

    for i, row in enumerate(errors):
        data = row["data"]
        flds = list(data.keys())
        if i == 0:
            ws.append(flds)
        ws.append(list(data.values()))

        row_errors = row["error"]

        for error in row_errors.errors():
            if len(error["loc"]):
                fld = error["loc"][0]
                idx = flds.index(fld)
            else:
                idx = 0
            cell = ws.cell(row=(i + 2), column=(idx + 1))
            cell.comment = Comment(error["msg"], "Schema Checker")
            cell.style = "Bad"

    wb.save(report_name)
